from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


OUT = Path(__file__).parent / "Semantic_Workspace_Template_v1.xlsx"


def style_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F2937")
    ws.freeze_panes = "A2"


def add_home_row(
    ws: Worksheet,
    word_uuid: str,
    verse_uuid: str,
    token_index: int,
    surface_text: str,
    lemma: str,
    strongs: str,
    pos: str,
    lang: str = "he",
) -> None:
    r = ws.max_row + 1
    ws.cell(r, 1, word_uuid)
    ws.cell(r, 2, verse_uuid)
    ws.cell(r, 3, "uuid_c_gen_1")
    ws.cell(r, 4, "uuid_b_genesis")
    ws.cell(r, 5, "uuid_t_ot")
    ws.cell(r, 6, token_index)
    ws.cell(r, 7, surface_text)
    ws.cell(r, 8, lemma)
    ws.cell(r, 9, strongs)
    ws.cell(r, 10, pos)
    ws.cell(r, 11, lang)

    ref_cell = ws.cell(r, 12, "Go -> References")
    ref_cell.hyperlink = f"#word_references!A1"
    ref_cell.style = "Hyperlink"

    act_cell = ws.cell(r, 13, "Go -> Actions")
    act_cell.hyperlink = f"#word_actions!A1"
    act_cell.style = "Hyperlink"

    note_cell = ws.cell(r, 14, "Go -> Notes")
    note_cell.hyperlink = f"#word_notes!A1"
    note_cell.style = "Hyperlink"


def autosize(ws: Worksheet) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)


def build() -> None:
    wb = Workbook()

    # 1) Home sheet: one row per word
    tokens = wb.active
    tokens.title = "tokens_home"
    style_header(
        tokens,
        [
            "word_uuid",
            "verse_uuid",
            "chapter_uuid",
            "book_uuid",
            "testament_uuid",
            "token_index",
            "surface_text",
            "lemma",
            "strongs",
            "pos",
            "lang",
            "references_link",
            "actions_link",
            "notes_link",
        ],
    )

    # Genesis 1:1 starter rows
    add_home_row(tokens, "uuid_w_gen_1_1_001", "uuid_v_gen_1_1", 1, "In", "be-", "H7225", "prep")
    add_home_row(tokens, "uuid_w_gen_1_1_002", "uuid_v_gen_1_1", 2, "the", "ha-", "", "article")
    add_home_row(tokens, "uuid_w_gen_1_1_003", "uuid_v_gen_1_1", 3, "beginning", "reshith", "H7225", "noun")
    add_home_row(tokens, "uuid_w_gen_1_1_004", "uuid_v_gen_1_1", 4, "God", "elohim", "H430", "noun")
    add_home_row(tokens, "uuid_w_gen_1_1_005", "uuid_v_gen_1_1", 5, "created", "bara", "H1254", "verb")

    # 2) References (definition/evidence/cross-ref/source)
    refs = wb.create_sheet("word_references")
    style_header(
        refs,
        [
            "ref_id",
            "word_uuid",
            "reference_type",
            "target_id",
            "label",
            "confidence",
            "status",
            "source_id",
            "notes",
        ],
    )
    refs.append(
        [
            "REF-0001",
            "uuid_w_gen_1_1_004",
            "definition",
            "DEF-ELOHIM",
            "Elohim lexical entry",
            0.95,
            "verified",
            "SRC-BDB",
            "Primary lexicon mapping",
        ]
    )
    refs.append(
        [
            "REF-0002",
            "uuid_w_gen_1_1_005",
            "crossref",
            "uuid_v_john_1_1",
            "Creation motif parallel",
            0.82,
            "provisional",
            "SRC-CROSSREF",
            "Needs review",
        ]
    )

    # 3) Actions (what user/system does with word)
    acts = wb.create_sheet("word_actions")
    style_header(
        acts,
        [
            "action_id",
            "word_uuid",
            "action_type",
            "action_payload",
            "owner",
            "status",
            "created_at",
            "notes",
        ],
    )
    acts.append(
        [
            "ACT-0001",
            "uuid_w_gen_1_1_005",
            "link_claim",
            "CL-A-001",
            "David",
            "open",
            "2026-03-09",
            "Link to creation-order claim",
        ]
    )

    # 4) Notes (freeform by word)
    notes = wb.create_sheet("word_notes")
    style_header(
        notes,
        [
            "note_id",
            "word_uuid",
            "note_type",
            "note_text",
            "author",
            "created_at",
            "updated_at",
        ],
    )
    notes.append(
        [
            "NOTE-0001",
            "uuid_w_gen_1_1_004",
            "theology",
            "Track plural-form discussion and semantic scope.",
            "David",
            "2026-03-09",
            "2026-03-09",
        ]
    )

    # 5+) Supporting evidence chain tabs
    claims = wb.create_sheet("claims")
    style_header(claims, ["claim_id", "claim_text", "claim_type", "status", "owner"])
    claims.append(["CL-A-001", "Creation is structured act, not random emergence.", "doctrinal", "draft", "David"])

    evidence = wb.create_sheet("evidence")
    style_header(evidence, ["evidence_id", "claim_id", "source_id", "excerpt", "score"])
    evidence.append(["EV-0001", "CL-A-001", "SRC-BDB", "Lexical root supports intentional creation action.", 0.88])

    sources = wb.create_sheet("sources")
    style_header(sources, ["source_id", "title", "author", "year", "url_or_path", "source_type"])
    sources.append(["SRC-BDB", "Brown-Driver-Briggs Hebrew Lexicon", "BDB", "1906", "", "lexicon"])
    sources.append(["SRC-CROSSREF", "Cross-reference notes", "Internal", "2026", "", "internal"])

    import_log = wb.create_sheet("import_log")
    style_header(import_log, ["import_id", "timestamp", "file_name", "rows_loaded", "errors"])

    instructions = wb.create_sheet("README")
    instructions["A1"] = "Semantic Workspace Excel Layer (v1)"
    instructions["A1"].font = Font(bold=True, size=14)
    instructions["A3"] = "Use tokens_home as the primary per-word table."
    instructions["A4"] = "Click links in columns L/M/N to jump to References, Actions, Notes."
    instructions["A5"] = "Keep UUIDs stable; never regenerate for existing rows."
    instructions["A6"] = "This workbook is import/export bridge for Obsidian + Forge."

    for ws in wb.worksheets:
        autosize(ws)

    wb.save(OUT)
    print(f"Wrote: {OUT}")


if __name__ == "__main__":
    build()
