from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


OUT = Path(__file__).parent / "Semantic_Workspace_Nodes_Edges_v2.xlsx"


def style_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F2937")
    ws.freeze_panes = "A2"


def autosize(ws: Worksheet) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 48)


def add_node(
    ws: Worksheet,
    node_id: str,
    short_id: str,
    title: str,
    statement: str,
    scope_level: str,
    entity_class: str,
    question_type: str,
    ops_status: str,
    chain_position: int,
    tree_level: int,
    origin_domain: str = "theophysics",
) -> None:
    r = ws.max_row + 1
    ws.cell(r, 1, node_id)
    ws.cell(r, 2, short_id)
    ws.cell(r, 3, title)
    ws.cell(r, 4, statement)
    ws.cell(r, 5, scope_level)
    ws.cell(r, 6, entity_class)
    ws.cell(r, 7, question_type)
    ws.cell(r, 8, ops_status)
    ws.cell(r, 9, chain_position)
    ws.cell(r, 10, tree_level)
    ws.cell(r, 11, origin_domain)
    ws.cell(r, 12, "2026-03-09T00:00:00Z")
    # support_score
    ws.cell(
        r,
        13,
        f'=SUMIFS(RELATIONSHIPS!$F:$F,RELATIONSHIPS!$B:$B,$A{r},RELATIONSHIPS!$D:$D,"evidence",RELATIONSHIPS!$E:$E,"supports")',
    )
    # attack_score
    ws.cell(
        r,
        14,
        f'=SUMIFS(RELATIONSHIPS!$F:$F,RELATIONSHIPS!$B:$B,$A{r},RELATIONSHIPS!$D:$D,"evidence",RELATIONSHIPS!$E:$E,"attacks")',
    )
    # structure_score
    ws.cell(
        r,
        15,
        f'=SUMIFS(RELATIONSHIPS!$F:$F,RELATIONSHIPS!$B:$B,$A{r},RELATIONSHIPS!$D:$D,"role")+SUMIFS(RELATIONSHIPS!$F:$F,RELATIONSHIPS!$B:$B,$A{r},RELATIONSHIPS!$D:$D,"time")+SUMIFS(RELATIONSHIPS!$F:$F,RELATIONSHIPS!$B:$B,$A{r},RELATIONSHIPS!$D:$D,"context")',
    )
    # death_penalty (0.20 per dead death condition)
    ws.cell(
        r,
        16,
        f'=COUNTIFS(RELATIONSHIPS!$B:$B,$A{r},RELATIONSHIPS!$G:$G,"<>none",RELATIONSHIPS!$H:$H,"dead")*0.20',
    )
    # raw_t_score
    ws.cell(r, 17, f"=O{r}+M{r}-N{r}-P{r}")
    # normalized 0-100
    ws.cell(r, 18, f"=MAX(0,MIN(100,ROUND(Q{r}*100,1)))")
    # tier
    ws.cell(r, 19, f'=IF(R{r}>=90,"near-canonical",IF(R{r}>=75,"strong",IF(R{r}>=60,"provisional","weak")))')


def add_edge(
    ws: Worksheet,
    edge_id: str,
    from_node_id: str,
    to_node_id: str,
    facet: str,
    edge_type: str,
    weight: float,
    death_condition: str = "none",
    status: str = "alive",
    source_ref: str = "",
    note: str = "",
) -> None:
    ws.append(
        [
            edge_id,
            from_node_id,
            to_node_id,
            facet,
            edge_type,
            weight,
            death_condition,
            status,
            source_ref,
            note,
            "2026-03-09T00:00:00Z",
        ]
    )


def build() -> None:
    wb = Workbook()

    nodes = wb.active
    nodes.title = "NODES"
    style_header(
        nodes,
        [
            "node_id",
            "short_id",
            "title",
            "statement",
            "scope_level",
            "entity_class",
            "question_type",
            "ops_status",
            "chain_position",
            "tree_level",
            "origin_domain",
            "created_at",
            "support_score",
            "attack_score",
            "structure_score",
            "death_penalty",
            "raw_t_score",
            "t_score",
            "tier",
        ],
    )

    rel = wb.create_sheet("RELATIONSHIPS")
    style_header(
        rel,
        [
            "edge_id",
            "from_node_id",
            "to_node_id",
            "facet",
            "edge_type",
            "weight",
            "death_condition",
            "status",
            "source_ref",
            "note",
            "created_at",
        ],
    )

    # Starter nodes
    add_node(
        nodes,
        "node-a1-1",
        "A1.1",
        "Existence Baseline",
        "Something exists rather than nothing.",
        "system",
        "axiom",
        "type1_binary",
        "canonical",
        1,
        0,
    )
    add_node(
        nodes,
        "node-a1-2",
        "A1.2",
        "Distinction Requirement",
        "Existence requires distinction.",
        "system",
        "axiom",
        "type2_identity",
        "reviewed",
        2,
        1,
    )
    add_node(
        nodes,
        "node-q0b",
        "Q0-B",
        "Anything Exists Branch",
        "Affirmation branch from Q0.",
        "system",
        "branch",
        "type1_binary",
        "reviewed",
        1,
        0,
    )

    # Starter relationships
    add_edge(rel, "edge-001", "node-a1-1", "node-a1-2", "time", "forces", 0.90, note="A1.1 forces A1.2")
    add_edge(rel, "edge-002", "node-a1-1", "node-a1-1", "evidence", "supports", 0.95, source_ref="self_refutation_test")
    add_edge(
        rel,
        "edge-003",
        "node-a1-1",
        "node-a1-1",
        "evidence",
        "attacks",
        0.20,
        death_condition="self_refutation",
        status="dead",
        source_ref="null_hypothesis",
        note="Negation path collapses",
    )
    add_edge(rel, "edge-004", "node-a1-1", "node-q0b", "role", "propagates_to", 0.85)
    add_edge(rel, "edge-005", "node-a1-1", "node-a1-1", "context", "maps_to", 0.70, note="maps to scripture/physics")

    readme = wb.create_sheet("README")
    readme["A1"] = "Semantic Workspace Nodes+Edges Template (v2)"
    readme["A1"].font = Font(bold=True, size=14)
    readme["A3"] = "This workbook has two canonical tabs: NODES and RELATIONSHIPS."
    readme["A4"] = "T-score is computed in NODES from weighted RELATIONSHIPS."
    readme["A5"] = "Edit edge weights/death conditions to update truth-confidence."
    readme["A6"] = "Everything else (tree view, bridge view, worldview view) is a filtered projection."
    readme["A8"] = "T-score formula:"
    readme["A9"] = "raw_t_score = structure_score + support_score - attack_score - death_penalty"
    readme["A10"] = "t_score = clamp(raw_t_score * 100, 0, 100)"
    readme["A11"] = "tier: near-canonical/strong/provisional/weak"

    for ws in wb.worksheets:
        autosize(ws)

    wb.save(OUT)
    print(f"Wrote: {OUT}")


if __name__ == "__main__":
    build()
